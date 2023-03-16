from django.shortcuts import render,redirect,HttpResponse
from app01 import models
from app01.utils.form import UserModelForm,MobileModelForm,MobileEditModelForm

import openpyxl
from openpyxl import load_workbook

# ------------------------靓号管理----------------------------
def mobile(request):
    if request.method == "GET":
        data = {}
        value = request.GET.get('num','')
        if value:
            data['mobile_' \
                 '_contains'] = value
        mobile_list = models.PrettyNum.objects.filter(**data)
        return render(request,'mobile.html',{"mobile_list":mobile_list,"value":value})


def mobile_add(request):
    if request.method == "GET":
        form = MobileModelForm()
        return render(request,'mobile_add.html',{"form":form})

    form = MobileModelForm(request.POST)
    if form.is_valid():
        form.save()
        form = MobileModelForm()
        flag = True
        return render(request,'mobile_add.html',{"form":form,"flag":flag})

    return render(request,'mobile_add.html',{"form":form})


def mobile_edit(request,nid):
    row_obj = models.PrettyNum.objects.filter(id=nid).first()

    if request.method == 'GET':
        form = MobileEditModelForm(instance=row_obj)
        return render(request,'mobile_edit.html',{'form':form})

    form = MobileEditModelForm(request.POST,instance=row_obj)
    if form.is_valid():
        form.save()
        flag = True
        return render(request,'mobile_edit.html',{"flag":flag,"form":form})

    return render(request,'mobile_edit.html',{"form":form})

def mobile_delete(request,nid):
    models.PrettyNum.objects.filter(id=nid).delete()
    return redirect('/mobile/')

# 使用文件上传功能批量添加数据
def mobile_multiadd(request):
    obj = request.FILES.get('excel')
    wb = load_workbook(obj)
    st = wb.worksheets[0]
    # cl = st.cell(2,2)
    data_list = []
    for row in st.iter_rows(min_row=2):
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        data_list.append(row_list)
    # print(data_list)
    for data in data_list:
        models.PrettyNum.objects.create(mobile=data[0],price=data[1],level=data[2],status=data[3])
        # 需要自行进行数据的校验


    return redirect('/mobile/')